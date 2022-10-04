import os
import csv
import pytest
from openpyxl import load_workbook
from PyPDF2 import PdfReader
from zipfile import ZipFile, ZIP_DEFLATED


pdf_file = 'docs-pytest-org-en-latest.pdf'
xlsx_file = 'file_example_XLSX_50.xlsx'
csv_file = 'SampleCSVFile_11kb.csv'
list_files = [pdf_file, xlsx_file, csv_file]
path = os.path.abspath('./resources')
zip_file = 'test.zip'


@pytest.mark.parametrize('path,zip_file,list_files', [[path, zip_file, list_files]])
def test_create_zip_file(path, zip_file, list_files):
    with ZipFile(os.path.join(path, zip_file), mode='w', compression=ZIP_DEFLATED) as zf:
        for file in list_files:
            zf.write(file)

    with ZipFile(os.path.join(path, zip_file), mode='a') as zf:
        for file in zf.namelist():
            assert file in list_files


@pytest.mark.parametrize('pdf_file', [pdf_file])
def test_check_pdf_file(pdf_file):
    with ZipFile(os.path.join(path, zip_file), mode='r') as zf:
            pdf_reader = PdfReader(zf.extract(pdf_file))
            number_of_pages = len(pdf_reader.pages)
            page_five = pdf_reader.pages[4].extract_text()
            assert number_of_pages == 412
            assert 'pytest Documentation, Release 0.1' in page_five


@pytest.mark.parametrize('xlsx_file', [xlsx_file])
def test_check_xlsx_file(xlsx_file):
    with ZipFile(os.path.join(path, zip_file), mode='r') as zf:
        workbook = load_workbook(zf.extract(xlsx_file))
        sheet = workbook.active
        assert sum((1 for row in sheet.values)) == 51
        assert sheet.cell(row=3, column=2).value == 'Mara'


@pytest.mark.parametrize('csv_file', [csv_file])
def test_check_csv_file(csv_file):
    with ZipFile(os.path.join(path, zip_file), mode='r') as zf:
        with open(zf.extract(csv_file)) as csvfile:
            table = csv.reader(csvfile)
            line_two = [line for line_no, line in enumerate(table, 1) if line_no == 2][0]
            assert line_two[2] == 'Barry French'
            assert table.line_num == 100


